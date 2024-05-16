import openpyxl
from openpyxl import load_workbook
import re
import PyPDF2
import pandas as pd #slow
import tkinter as tk #slow
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk # For dropdown menu
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import skrf as rf
import os
import numpy as np
import scipy
import control as ct  # control systems tools

"""
    Summary:
    Tools for SNP files ex. plotting, converting to excel files, various reports coming soon
    
    Potential Additions:
    - Multiple files at once
    - Choosing between linear and log scales
    - Choosing between subplot and over-plotting
    - Adding features for PIM, EMC etc tests
    - Save plot photos
    - Generate new sheet/file for excel output
    - Error handling
    - Improve GUI
    - Package as runnable .exe for sharing with the team
    - GIT?
"""




# Function to show the selected frame
def show_frame(frame):
    frame.tkraise()


# Add content to landing frame
def on_option_selected(event):
    selected_option = dropdown.get()
    if selected_option == "SNP":
        show_frame(snpFrame)
    elif selected_option == "SOC":
        show_frame(socFrame)


# Define click events for GUI
def browse_Excel():
    dir_path = filedialog.askopenfilename()
    if dir_path:
        entry_dir1.delete(0, tk.END)
        entry_dir1.insert(0, dir_path)

def browse_SNP():
    dir_path = filedialog.askopenfilename()
    if dir_path:
        entry_dir2.delete(0, tk.END)
        entry_dir2.insert(0, dir_path)

def show_error(message):
    messagebox.showerror("Error", message)

def submit():
    excelPath = entry_dir1.get()
    snpPath = entry_dir2.get()
    sweepData = rf.Network(snpPath)
    if excelPath:
        convert2Excel(sweepData, excelPath)
    if not snpPath:
        show_error("Please enter an SNP path and resubmit")

    button_submit.config(bg="#90EE90")

    sParams = ct.mag2db(np.abs(sweepData.s))
    numPorts = sweepData.number_of_ports

    fig, ax = plt.subplots()
    for port1 in range(numPorts):
        for port2 in range(numPorts):
            ax.plot(sweepData.f, sParams[:, port1, port2], label=f"S{port1+1}-{port2+1}")

    ax.set_xlabel('Frequency (Hz)')
    ax.set_ylabel('Magnitude (dB)')
    ax.set_title('S-Parameters vs Frequency')
    ax.legend()
    ax.grid(True)

    canvas = FigureCanvasTkAgg(fig, master=mainWindow)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=1, column=2, rowspan=3, padx=10, pady=10, sticky="nsew")

def RemoveExcelData(filePath):
    wb = openpyxl.load_workbook(filePath)
    ws = wb['VNA Output']
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    wb.save(filePath)

def convert2Excel(data, filePath):
    sParams = ct.mag2db(np.abs(data.s))
    numPorts = data.number_of_ports
    wb = openpyxl.load_workbook(filePath)
    sheet = wb["VNA Output"]
    for freq, value in enumerate(data.f):
        if freq == 0:
            sheet.cell(row=1, column=1).value = "Frequency"
        sheet.cell(row=freq+2, column=1).value = data.f[freq]
        for port1 in range(numPorts):
            for port2 in range(numPorts):
                if freq == 0:
                    sheet.cell(row=1, column=2 + port1*numPorts + port2).value = f"S{port1+1}-{port2+1} (Mag dB)"
                sheet.cell(row=freq+2, column=2 + port1*numPorts + port2).value = sParams[freq, port1, port2]
    wb.save(filePath)
    
    
    
# Create a Tkinter windows
mainWindow = tk.Tk()
mainWindow.title("RF Tools")

# Define frames for different pages
landingFrame = tk.Frame(mainWindow)
snpFrame = tk.Frame(mainWindow)
socFrame = tk.Frame(mainWindow)
# Set up the landing page with dropdown menu
landingFrame.grid(row=0, column=0, sticky="nsew")
snpFrame.grid(row=0, column=0, sticky="nsew")

landingLabel = tk.Label(landingFrame, text="Select an Option", font=("Humanist 777", 16, "bold"))
landingLabel.pack(pady=20)

options = ["SNP", "SOC"]
dropdown = ttk.Combobox(landingFrame, values=options)
dropdown.bind("<<ComboboxSelected>>", on_option_selected)
dropdown.pack(pady=20)

# Add content to excel frame
button_browse1 = tk.Button(snpFrame, text="Select Excel file", command=browse_Excel, activebackground='#90EE90', bg='#FFFFFF', font=("Humanist 777", 10, "bold"), fg='#0072C6')
button_browse1.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

button_browse2 = tk.Button(snpFrame, text="Select SNP file", command=browse_SNP, activebackground='#90EE90', bg='#FFFFFF', font=("Humanist 777", 10, "bold"), fg='#0072C6')
button_browse2.grid(row=2, column=0, padx=10, pady=5, sticky="ew")

button_submit = tk.Button(snpFrame, text="Submit", command=submit, bg='#FFFFFF', font=("Humanist 777", 10, "bold"), fg='#0072C6')
button_submit.grid(row=3, column=0, padx=10, pady=5, sticky="ew")

entry_dir1 = tk.Entry(snpFrame)
entry_dir1.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
entry_dir2 = tk.Entry(snpFrame)
entry_dir2.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

for i in range(4):
    snpFrame.grid_rowconfigure(i, weight=1)
mainWindow.grid_columnconfigure(2, weight=1)

# Show landing frame initially
show_frame(landingFrame)

mainWindow.config(bg="#D3D3D3")
screen_width = mainWindow.winfo_screenwidth()
screen_height = mainWindow.winfo_screenheight()
mainWindow.geometry(f"{screen_width}x{screen_height}")
mainWindow.mainloop()
